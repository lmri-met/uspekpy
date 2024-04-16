import math
import os
from pathlib import Path
from time import time

import numpy as np
import openpyxl
import pandas as pd
import spekpy as sp
from scipy import interpolate

import uspeckpy.tools as tl


def main(beam_data_file, conversion_coefficients_files, transmission_coefficients_file,
         transmission_coefficients_uncertainty, simulations_number, output_folder):
    print('RUNNING')
    # PAL code
    lista_mono = []

    # XCB: INPUT DATA DIGEST: SIMULATIONS NUMBER, UNCERTAINTY MU_TR, OUTPUT FOLDER, CONVERSION COEFFICIENTS FILES
    # ------------------------------------------------------------------------------------------------------------------
    # XCB code
    nini = simulations_number
    umutrr = transmission_coefficients_uncertainty
    ruta = output_folder
    directorio = ''
    for x in conversion_coefficients_files:
        lista_mono.append(os.path.basename(x))
        directorio = os.path.dirname(x)
    # XCB: END OF INPUT DATA DIGEST: SIMULATIONS NUMBER, UNCERTAINTY MU_TR, OUTPUT FOLDER, CONVERSION COEFFICIENTS FILES
    # ------------------------------------------------------------------------------------------------------------------

    # XCB: INPUT DATA DIGEST: BEAM DATA FILE
    # ------------------------------------------------------------------------------------------------------------------
    # PAL code
    for x in range(2, 47):
        # Opens Excel file
        archivo_excel = openpyxl.load_workbook(beam_data_file)
        # selects working sheet by default
        hoja = archivo_excel.active
        # Acceds to cell B1 (row 1, column 2...47)
        z = 2
        calidad = hoja.cell(row=1, column=x).value
        # XCB: Add condidtion to break the loop if calidad is None
        if calidad is None:
            break
        # XCB: Keep reading if calidad is not None
        filter_Al = hoja.cell(row=z, column=x).value
        z = z + 1
        filter_Cu = hoja.cell(row=z, column=x).value
        z = z + 1
        filter_Sn = hoja.cell(row=z, column=x).value
        z = z + 1
        filter_Pb = hoja.cell(row=z, column=x).value
        z = z + 1
        filter_Be = hoja.cell(row=z, column=x).value
        z = z + 1
        filter_Air = hoja.cell(row=z, column=x).value
        z = z + 1
        kvp = hoja.cell(row=z, column=x).value
        z = z + 1
        th = hoja.cell(row=z, column=x).value
        z = z + 1
        uAl = hoja.cell(row=z, column=x).value
        z = z + 1
        uCu = hoja.cell(row=z, column=x).value
        z = z + 1
        uSn = hoja.cell(row=z, column=x).value
        z = z + 1
        uPb = hoja.cell(row=z, column=x).value
        z = z + 1
        uBe = hoja.cell(row=z, column=x).value
        z = z + 1
        uAir = hoja.cell(row=z, column=x).value
        z = z + 1
        ukvp = hoja.cell(row=z, column=x).value
        z = z + 1
        uth = hoja.cell(row=z, column=x).value
        # XCB: END OF INPUT DATA DIGEST: BEAM DATA FILE
        # ------------------------------------------------------------------------------------------------------------------

        lower_be = filter_Be * (1 - uBe * math.sqrt(3))
        high_be = filter_Be * (1 + uBe * math.sqrt(3))

        lower_al = filter_Al * (1 - uAl * math.sqrt(3))
        high_al = filter_Al * (1 + uAl * math.sqrt(3))

        lower_cu = filter_Cu * (1 - uCu * math.sqrt(3))
        high_cu = filter_Cu * (1 + uCu * math.sqrt(3))

        lower_sn = filter_Sn * (1 - uSn * math.sqrt(3))
        high_sn = filter_Sn * (1 + uSn * math.sqrt(3))

        lower_pb = filter_Pb * (1 - uPb * math.sqrt(3))
        high_pb = filter_Pb * (1 + uPb * math.sqrt(3))

        # definition of empty lists to calculate std deviation and std. rel uncertainty, ur% (variation coefficient %)
        energiasMedia = []
        kermasMedia = []
        hvlmean = []
        hvl2mean = []
        hvlCumean = []
        hvl2Cumean = []
        hpkMedia = []
        hpkMedia15 = []
        hpkMedia30 = []
        hpkMedia45 = []
        hpkMedia60 = []
        hpkMedia75 = []
        hpkMedia90 = []
        hpkMedia180 = []

        # XCB: INPUT DATA DIGEST: OUTPUT FOLDER, CONVERSION COEFFICIENTS FILES
        # ------------------------------------------------------------------------------------------------------------------

        # CREATE DIRECTORY TO STORE OUTPUT FILES
        Path(ruta).mkdir(parents=True, exist_ok=True)

        # ARRAY WITH SPECTRA FILE NAMES AND MONOENERGETIC CONV. COEFF.
        # PAL: list of monoenergetic coefficients (8 different depending on angles)
        ficheros_monoenergeticos = lista_mono

        # READ MONOENERGETIC FILES
        for f_m in ficheros_monoenergeticos:
            hk_table = pd.read_csv(directorio + "/" + f_m, sep=";", encoding='ISO-8859-1')
            tl.buscar_eliminar(ruta + "/" + f_m.upper() + ".txt")
            tl.buscar_eliminar(ruta + "/" + f_m.upper() + "_0.txt")
            tl.buscar_eliminar(ruta + "/" + f_m.upper() + "_15.txt")
            tl.buscar_eliminar(ruta + "/" + f_m.upper() + "_30.txt")
            tl.buscar_eliminar(ruta + "/" + f_m.upper() + "_45.txt")
            tl.buscar_eliminar(ruta + "/" + f_m.upper() + "_60.txt")
            tl.buscar_eliminar(ruta + "/" + f_m.upper() + "_75.txt")
            tl.buscar_eliminar(ruta + "/" + f_m.upper() + "_90.txt")
            tl.buscar_eliminar(ruta + "/" + f_m.upper() + "_180.txt")
            # XCB: END OF INPUT DATA DIGEST: OUTPUT FOLDER, CONVERSION COEFFICIENTS FILES
            # ----------------------------------------------------------------------------------------------------------

            # FIRST LOOP: monoenergetic conversion coefficients defined at incident angle = 0
            # ("h_amb_10.csv","hp_0.07_pill.csv", "hp_0.07_rod.csv")
            # ----------------------------------------------------------------------------------------------------------
            # TEST THAT THE FILE HAS LESS OR EQUAL THAN 2 COLUMNS
            if len(hk_table.columns) <= 2:
                print('FIRST LOOP: monoenergetic conversion coefficients defined at incident angle = 0')
                tiempo_inicial_columns_2 = time()

                # SAVE VARIABLES (E, hK)
                Ehk = hk_table.iloc[:, 0].values
                hk = hk_table.iloc[:, 1].values

                # take logs for interpolation
                LEhk = []
                Lhk = []

                for x in Ehk:
                    LEhk.append(math.log(x))

                for x in hk:
                    Lhk.append(math.log(x))

                # READ FILE MUTR_RHO
                mutr_rho = pd.read_csv(transmission_coefficients_file, sep="\t", header=None, encoding='ISO-8859-1')
                # SAVE VARIABLES (E, mutr/rho)
                E_mut_aux = mutr_rho.iloc[:, 0].values
                mtr_aux = mutr_rho.iloc[:, 1].values

                E_mut = []
                mtr = []

                # REPLACE IF NEEDED COMMAS AND DOTS
                for x in E_mut_aux:
                    E_mut.append(float(x))

                for x in mtr_aux:
                    mtr.append(float(x))

                LE_mut = [math.log(x) for x in E_mut]
                Lmtr = [math.log(x) for x in mtr]

                # NINI = Number of times that a spectrum is randomly generated (the parameters that define
                #   a quality are changed around a central value using a gaussian or uniform distribution
                for j in range(nini):
                    if filter_Al != 0 and uAl != 0:
                        filter_Al_rand = np.random.uniform(lower_al, high_al, 1)
                    elif filter_Al != 0 and uAl == 0 and j == 0:
                        filter_Al_rand = np.random.uniform(lower_al, high_al, 1)
                    elif filter_Al == 0 and j == 0:
                        filter_Al_rand = np.random.uniform(lower_al, high_al, 1)

                    if filter_Be != 0 and uBe != 0:
                        filter_Be_rand = np.random.uniform(lower_be, high_be, 1)
                    elif filter_Be != 0 and uBe == 0 and j == 0:
                        filter_Be_rand = np.random.uniform(lower_be, high_be, 1)
                    elif filter_Be == 0 and j == 0:
                        filter_Be_rand = np.random.uniform(lower_be, high_be, 1)

                    if filter_Cu != 0 and uCu != 0:
                        filter_Cu_rand = np.random.uniform(lower_cu, high_cu, 1)
                    elif filter_Cu != 0 and uCu == 0 and j == 0:
                        filter_Cu_rand = np.random.uniform(lower_cu, high_cu, 1)
                    elif filter_Cu == 0 and j == 0:
                        filter_Cu_rand = np.random.uniform(lower_cu, high_cu, 1)

                    if filter_Sn != 0 and uSn != 0:
                        filter_Sn_rand = np.random.uniform(lower_sn, high_sn, 1)
                    elif filter_Sn != 0 and uSn == 0 and j == 0:
                        filter_Sn_rand = np.random.uniform(lower_sn, high_sn, 1)
                    elif filter_Sn == 0 and j == 0:
                        filter_Sn_rand = np.random.uniform(lower_sn, high_sn, 1)

                    if filter_Pb != 0 and uPb != 0:
                        filter_Pb_rand = np.random.uniform(lower_pb, high_pb, 1)
                    elif filter_Pb != 0 and uPb == 0 and j == 0:
                        filter_Pb_rand = np.random.uniform(lower_pb, high_pb, 1)
                    elif filter_Pb == 0 and j == 0:
                        filter_Pb_rand = np.random.uniform(lower_pb, high_pb, 1)

                    if filter_Air != 0 and uAir != 0:
                        filter_Air_rand = np.random.normal(filter_Air, uAir * filter_Air, 1)
                    elif filter_Air != 0 and uAir == 0 and j == 0:
                        filter_Air_rand = np.random.normal(filter_Air, uAir * filter_Air, 1)
                    elif filter_Air == 0 and j == 0:
                        filter_Air_rand = np.random.normal(filter_Air, uAir * filter_Air, 1)

                    if kvp != 0 and ukvp != 0:
                        kvp_rand = np.random.normal(kvp, ukvp * kvp, 1)
                    elif kvp != 0 and ukvp == 0 and j == 0:
                        kvp_rand = np.random.normal(kvp, ukvp * kvp, 1)
                    elif kvp == 0 and j == 0:
                        kvp_rand = np.random.normal(kvp, ukvp * kvp, 1)

                    if th != 0 and uth != 0:
                        th_rand = np.random.normal(th, uth * th, 1)
                    elif th != 0 and uth == 0 and j == 0:
                        th_rand = np.random.normal(th, uth * th, 1)
                    elif th_rand == 0 and j == 0:
                        th_rand = np.random.normal(th, uth * th, 1)

                    # SPEKPY
                    my_filters = [
                        ("Al", filter_Al_rand[0]),
                        ("Be", filter_Be_rand[0]),
                        ("Air", filter_Air_rand[0]),
                        ("Cu", filter_Cu_rand[0]),
                        ("Sn", filter_Sn_rand[0]),
                        ("Pb", filter_Pb_rand[0]),
                    ]

                    s = sp.Spek(kvp=kvp_rand[0], th=th_rand[0])
                    s.multi_filter(my_filters)

                    E, fluencia = s.get_spectrum(edges=False)  # Get the spectrum
                    hvl = s.get_hvl1()  # Get the 1st half-value-layer mm Al
                    hvlmean.append(hvl)
                    hvl2 = s.get_hvl2()  # Get the 2nd half-value-layer mm Al
                    hvl2mean.append(hvl2)

                    hvlCu = s.get_hvl1(matl='Cu')  # Get the 1st half-value-layer mm Cu
                    hvl2Cu = s.get_hvl2(matl='Cu')  # Get the 2nd half-value-layer mm Cu
                    hvlCumean.append(hvlCu)
                    hvl2Cumean.append(hvl2Cu)
                    # END SPEKPY

                    # FILTERING SPECTRUM TO AVOID INTREPOLATION ERRORS (monoenergetic hK start sometimes at 7 keV)
                    mascara = E >= 8
                    E_filtrado_temp = E[mascara]
                    fluencia_filtrada_temp = fluencia[mascara]

                    # Calculation of mean energy
                    # first creating 2 empty lists to acumulate data from all loop iterations
                    fluencias_acumuladas = []
                    energias_acumuladas = []

                    # Adding values to empty lists
                    fluencias_acumuladas.extend(fluencia_filtrada_temp)
                    energias_acumuladas.extend(E_filtrado_temp)
                    # plt.plot(E_filtrado_temp, fluencia_filtrada_temp, label=f"Iteración {j+1}")
                    # plt.plot(E_filtrado_temp, fluencia_filtrada_temp)

                    # Calculation of mean energy out of the loop
                    energia_media = tl.calcular_energia_media(fluencias_acumuladas, energias_acumuladas)

                    energiasMedia.append(energia_media)

                    # Take log
                    LE = []
                    for x in E_filtrado_temp:
                        LE.append(math.log(x))

                    # Making Akima interpolation with MUTR_RHO (after taking logs)
                    interpolacion_uno_mut = interpolate.Akima1DInterpolator(LE_mut, Lmtr, axis=0)
                    list_mutr = []

                    for i in LE:
                        mutrrho = math.exp(interpolacion_uno_mut(i))
                        mutr = np.random.normal(mutrrho, umutrr * mutrrho,
                                                1)  # HERE mutr/rho is sampled randomly from gaussian distrib
                        list_mutr.append(mutr)

                    coeficientes_mutr_rho = []
                    coeficientes_mutr_rho.extend(list_mutr)

                    # Calculation of mean air kerma out of the loop
                    kerma_medio = tl.calcular_kerma_medio(fluencias_acumuladas, energias_acumuladas,
                                                          coeficientes_mutr_rho)
                    kermasMedia.append(kerma_medio)

                    # Making Akima interpolation with monoenergetic hK (after taking logs)
                    interpolacion_uno_hk = interpolate.Akima1DInterpolator(LEhk, Lhk, axis=0)
                    ln_hk_int = []

                    for i in LE:
                        interpolacion_final_hk = math.exp(interpolacion_uno_hk(i))
                        ln_hk_int.append(interpolacion_final_hk)

                    hk_int = ln_hk_int
                    hk = []
                    hk.extend(hk_int)

                    # Calculation of mean conv. coeff out of the loop
                    hpk = tl.calcular_hpk(fluencias_acumuladas, energias_acumuladas, coeficientes_mutr_rho, hk)
                    hpkMedia.append(hpk)

                np.set_printoptions(linewidth=np.inf)

                # MEAN ENERGY
                media_energia = np.mean(energiasMedia)
                desviacion_energia = np.std(energiasMedia, ddof=0)
                coef_variacion_energia = (desviacion_energia / media_energia) * 100

                # MEAN AIR KERMA
                media_kerma = np.mean(kermasMedia)
                desviacion_kerma = np.std(kermasMedia, ddof=0)
                coef_variacion_kerma = (desviacion_kerma / media_kerma) * 100

                # MEAN CONVERSION COEFFICIENT
                media_hpk = np.mean(hpkMedia)
                sd_hpk = np.std(hpkMedia, ddof=0)
                v_hpk = (sd_hpk / media_hpk) * 100

                # MEAN 1st HVL
                mean_hvl = np.mean(hvlmean)
                sd_hvl = np.std(hvlmean, ddof=0)
                v_hvl = (sd_hvl / mean_hvl) * 100

                mean_hvlCu = np.mean(hvlCumean)
                sd_hvlCu = np.std(hvlCumean, ddof=0)
                v_hvlCu = (sd_hvlCu / mean_hvl) * 100

                # MEAN 2nd HVL
                mean_hvl2 = np.mean(hvl2mean)
                sd_hvl2 = np.std(hvl2mean, ddof=0)
                v_hvl2 = (sd_hvl2 / mean_hvl2) * 100

                mean_hvl2Cu = np.mean(hvl2Cumean)
                sd_hvl2Cu = np.std(hvl2Cumean, ddof=0)
                v_hvl2Cu = (sd_hvl2Cu / mean_hvl2Cu) * 100

                tiempo_final_columns_2 = time()

                tiempo_ejecucion_columns_2 = tiempo_final_columns_2 - tiempo_inicial_columns_2

                # XCB: OUTPUT DATA DIGEST
                # ------------------------------------------------------------------------------------------------------
                # Lista de ángulos y sus respectivos valores hpk y sd_hpk
                angulos = [(0, float(hpk), sd_hpk, v_hpk)]

                # List of magnitudes independent of the angle
                magnitudes = [(media_energia, desviacion_energia, coef_variacion_energia),
                              (media_kerma, desviacion_kerma, coef_variacion_kerma),
                              (mean_hvl, sd_hvl, v_hvl),
                              (mean_hvl2, sd_hvl2, v_hvl2),
                              (mean_hvlCu, sd_hvlCu, v_hvlCu),
                              (mean_hvl2Cu, sd_hvl2Cu, v_hvl2Cu)]

                # Save results in TXT and CSV files
                df = tl.output_digest(operational_magnitude=f_m, quality=calidad, mean_magnitudes=magnitudes,
                                      conversion_coefficients=angulos, execution_time=tiempo_ejecucion_columns_2,
                                      output_folder=ruta)

                print('Execution time in (s):', tiempo_ejecucion_columns_2)

            # SECOND LOOP: monoenergetic coefficients defined at 6 incident angles: 0 - 75 deg
            # ("hp_10_slab.csv","hp_0.07_slab.csv")
            # ----------------------------------------------------------------------------------------------------------
            elif len(hk_table.columns) == 7:
                print('SECOND LOOP: monoenergetic coefficients defined at 6 incident angles: 0 - 75 deg')
                tiempo_inicial_columns_7 = time()

                # SLICING THE TABLES WHEN DISCOVERING ZEROES.THIS AVOIDS CRASHING WHEN TAKING LOGS AND DOING AKIMA
                hk_table15 = hk_table[hk_table['15angle'] > 0]
                hk_table30 = hk_table[hk_table['30angle'] > 0]
                hk_table45 = hk_table[hk_table['45angle'] > 0]
                hk_table60 = hk_table[hk_table['60angle'] > 0]
                hk_table75 = hk_table[hk_table['75angle'] > 0]

                Ehk = hk_table.iloc[:, 0].values
                hk_0 = hk_table.iloc[:, 1].values

                Ehk15 = hk_table15.iloc[:, 0].values
                hk_15 = hk_table15.iloc[:, 2].values

                Ehk30 = hk_table30.iloc[:, 0].values
                hk_30 = hk_table30.iloc[:, 3].values

                Ehk45 = hk_table45.iloc[:, 0].values
                hk_45 = hk_table45.iloc[:, 4].values

                Ehk60 = hk_table60.iloc[:, 0].values
                hk_60 = hk_table60.iloc[:, 5].values

                Ehk75 = hk_table75.iloc[:, 0].values
                hk_75 = hk_table75.iloc[:, 6].values

                # TAKE LOGS FOR FURTHER AKIMA INTERPOLATIONS
                LEhk = []
                LEhk15 = []
                LEhk30 = []
                LEhk45 = []
                LEhk60 = []
                LEhk75 = []
                Lhk_0 = []
                Lhk_15 = []
                Lhk_30 = []
                Lhk_45 = []
                Lhk_60 = []
                Lhk_75 = []

                [LEhk.append(math.log(x)) for x in Ehk]
                [LEhk15.append(math.log(x)) for x in Ehk15]
                [LEhk30.append(math.log(x)) for x in Ehk30]
                [LEhk45.append(math.log(x)) for x in Ehk45]
                [LEhk60.append(math.log(x)) for x in Ehk60]
                [LEhk75.append(math.log(x)) for x in Ehk75]

                [Lhk_0.append(math.log(x)) for x in hk_0]
                [Lhk_15.append(math.log(x)) for x in hk_15]
                [Lhk_30.append(math.log(x)) for x in hk_30]
                [Lhk_45.append(math.log(x)) for x in hk_45]
                [Lhk_60.append(math.log(x)) for x in hk_60]
                [Lhk_75.append(math.log(x)) for x in hk_75]

                # READ MUTR_RHO FILE
                mutr_rho = pd.read_csv(transmission_coefficients_file, sep="\t", header=None, encoding='ISO-8859-1')

                # SAVE VARIABLES MUTR_RHO vs Energy
                E_mut_aux = mutr_rho.iloc[:, 0].values
                mtr_aux = mutr_rho.iloc[:, 1].values

                E_mut = []
                mtr = []

                # Replace dots by commas
                for x in E_mut_aux:
                    E_mut.append(float(x))

                for x in mtr_aux:
                    mtr.append(float(x))

                # Take logs
                LE_mut = [math.log(x) for x in E_mut]
                Lmtr = [math.log(x) for x in mtr]

                # NINI = Number of times that a random spectrum is generated (the parameters that define
                # a quality are changed around a central value using a gaussian or uniform distribution
                for j in range(nini):
                    if filter_Al != 0 and uAl != 0:
                        filter_Al_rand = np.random.uniform(lower_al, high_al, 1)
                    elif filter_Al != 0 and uAl == 0 and j == 0:
                        filter_Al_rand = np.random.uniform(lower_al, high_al, 1)
                    elif filter_Al == 0 and j == 0:
                        filter_Al_rand = np.random.uniform(lower_al, high_al, 1)

                    if filter_Be != 0 and uBe != 0:
                        filter_Be_rand = np.random.uniform(lower_be, high_be, 1)
                    elif filter_Be != 0 and uBe == 0 and j == 0:
                        filter_Be_rand = np.random.uniform(lower_be, high_be, 1)
                    elif filter_Be == 0 and j == 0:
                        filter_Be_rand = np.random.uniform(lower_be, high_be, 1)

                    if filter_Cu != 0 and uCu != 0:
                        filter_Cu_rand = np.random.uniform(lower_cu, high_cu, 1)
                    elif filter_Cu != 0 and uCu == 0 and j == 0:
                        filter_Cu_rand = np.random.uniform(lower_cu, high_cu, 1)
                    elif filter_Cu == 0 and j == 0:
                        filter_Cu_rand = np.random.uniform(lower_cu, high_cu, 1)

                    if filter_Sn != 0 and uSn != 0:
                        filter_Sn_rand = np.random.uniform(lower_sn, high_sn, 1)
                    elif filter_Sn != 0 and uSn == 0 and j == 0:
                        filter_Sn_rand = np.random.uniform(lower_sn, high_sn, 1)
                    elif filter_Sn == 0 and j == 0:
                        filter_Sn_rand = np.random.uniform(lower_sn, high_sn, 1)

                    if filter_Pb != 0 and uPb != 0:
                        filter_Pb_rand = np.random.uniform(lower_pb, high_pb, 1)
                    elif filter_Pb != 0 and uPb == 0 and j == 0:
                        filter_Pb_rand = np.random.uniform(lower_pb, high_pb, 1)
                    elif filter_Pb == 0 and j == 0:
                        filter_Pb_rand = np.random.uniform(lower_pb, high_pb, 1)

                    if filter_Air != 0 and uAir != 0:
                        filter_Air_rand = np.random.normal(filter_Air, uAir * filter_Air, 1)
                    elif filter_Air != 0 and uAir == 0 and j == 0:
                        filter_Air_rand = np.random.normal(filter_Air, uAir * filter_Air, 1)
                    elif filter_Air == 0 and j == 0:
                        filter_Air_rand = np.random.normal(filter_Air, uAir * filter_Air, 1)

                    if kvp != 0 and ukvp != 0:
                        kvp_rand = np.random.normal(kvp, ukvp * kvp, 1)
                    elif kvp != 0 and ukvp == 0 and j == 0:
                        kvp_rand = np.random.normal(kvp, ukvp * kvp, 1)
                    elif kvp == 0 and j == 0:
                        kvp_rand = np.random.normal(kvp, ukvp * kvp, 1)

                    if th != 0 and uth != 0:
                        th_rand = np.random.normal(th, uth * th, 1)
                    elif th != 0 and uth == 0 and j == 0:
                        th_rand = np.random.normal(th, uth * th, 1)
                    elif th_rand == 0 and j == 0:
                        th_rand = np.random.normal(th, uth * th, 1)

                    # SPEKPY
                    my_filters = [
                        ("Al", filter_Al_rand[0]),
                        ("Be", filter_Be_rand[0]),
                        ("Air", filter_Air_rand[0]),
                        ("Cu", filter_Cu_rand[0]),
                        ("Sn", filter_Sn_rand[0]),
                        ("Pb", filter_Pb_rand[0]),
                    ]

                    s = sp.Spek(kvp=kvp_rand[0], th=th_rand[0])
                    s.multi_filter(my_filters)

                    E, fluencia = s.get_spectrum(edges=False)  # Obtener el espectro
                    hvl = s.get_hvl1()  # Get the 1st half-value-layer mm Al
                    hvlmean.append(hvl)
                    hvl2 = s.get_hvl2()  # Get the 2nd half-value-layer mm Al
                    hvl2mean.append(hvl2)

                    hvlCu = s.get_hvl1(matl='Cu')  # Get the 1st half-value-layer mm Cu
                    hvl2Cu = s.get_hvl2(matl='Cu')  # Get the 2nd half-value-layer mm Cu
                    hvlCumean.append(hvlCu)
                    hvl2Cumean.append(hvl2Cu)
                    # END SPEKPY

                    # FILTRADO DE SPEKPY Y OPERACIONES SOBRE ESPECTRO
                    # cortamos para que no de error al interpolar los hK monoenergetico (comienzan a veces en 7keV)
                    mascara = E >= 8
                    E_filtrado_temp = E[mascara]
                    fluencia_filtrada_temp = fluencia[mascara]

                    # Calculo de energia media
                    # primero debemos crear 2 listas vacias para que vaya acumuludando los datos de las iteraciones
                    fluencias_acumuladas = []
                    energias_acumuladas = []

                    # Agrego valores a las listas vacias
                    fluencias_acumuladas.extend(fluencia_filtrada_temp)
                    energias_acumuladas.extend(E_filtrado_temp)
                    energia_media = tl.calcular_energia_media(fluencias_acumuladas, energias_acumuladas)

                    energiasMedia.append(energia_media)

                    # Realizar EL LOGARITMO
                    LE = []
                    for x in E_filtrado_temp:
                        LE.append(math.log(x))

                    # Making Akima interpolation with MUTR_RHO (after taking logs)
                    interpolacion_uno_mut = interpolate.Akima1DInterpolator(LE_mut, Lmtr, axis=0)
                    list_mutr = []

                    for i in LE:
                        mutrrho = math.exp(interpolacion_uno_mut(i))
                        mutr = np.random.normal(mutrrho, umutrr * mutrrho,
                                                1)  # HERE mutr/rho is sampled randomly from gaussian distrib
                        list_mutr.append(mutr)

                    coeficientes_mutr_rho = []
                    coeficientes_mutr_rho.extend(list_mutr)

                    # Calculo de kerma medio fuera del bucle
                    kerma_medio = tl.calcular_kerma_medio(fluencias_acumuladas, energias_acumuladas,
                                                          coeficientes_mutr_rho)
                    kermasMedia.append(kerma_medio)

                    # Making Akima interpolation with monoenergetic hK (after taking logs)
                    interpolacion_uno_0 = interpolate.Akima1DInterpolator(LEhk, Lhk_0, axis=0)
                    interpolacion_uno_15 = interpolate.Akima1DInterpolator(LEhk15, Lhk_15, axis=0)
                    interpolacion_uno_30 = interpolate.Akima1DInterpolator(LEhk30, Lhk_30, axis=0)
                    interpolacion_uno_45 = interpolate.Akima1DInterpolator(LEhk45, Lhk_45, axis=0)
                    interpolacion_uno_60 = interpolate.Akima1DInterpolator(LEhk60, Lhk_60, axis=0)
                    interpolacion_uno_75 = interpolate.Akima1DInterpolator(LEhk75, Lhk_75, axis=0)

                    ln_hk_int_0 = []
                    ln_hk_int_15 = []
                    ln_hk_int_30 = []
                    ln_hk_int_45 = []
                    ln_hk_int_60 = []
                    ln_hk_int_75 = []

                    for i in LE:
                        interpolacion_final_0 = math.exp(interpolacion_uno_0(i))
                        ln_hk_int_0.append(interpolacion_final_0)

                        interpolacion_final_15 = math.exp(interpolacion_uno_15(i))
                        ln_hk_int_15.append(interpolacion_final_15)

                        interpolacion_final_30 = math.exp(interpolacion_uno_30(i))
                        ln_hk_int_30.append(interpolacion_final_30)

                        interpolacion_final_45 = math.exp(interpolacion_uno_45(i))
                        ln_hk_int_45.append(interpolacion_final_45)

                        interpolacion_final_60 = math.exp(interpolacion_uno_60(i))
                        ln_hk_int_60.append(interpolacion_final_60)

                        interpolacion_final_75 = math.exp(interpolacion_uno_75(i))
                        ln_hk_int_75.append(interpolacion_final_75)

                    hk_int_0 = ln_hk_int_0
                    hk_int_15 = ln_hk_int_15
                    hk_int_30 = ln_hk_int_30
                    hk_int_45 = ln_hk_int_45
                    hk_int_60 = ln_hk_int_60
                    hk_int_75 = ln_hk_int_75

                    hk = []
                    hk15 = []
                    hk30 = []
                    hk45 = []
                    hk60 = []
                    hk75 = []

                    hk.extend(hk_int_0)
                    hpk = tl.calcular_hpk(fluencias_acumuladas, energias_acumuladas, coeficientes_mutr_rho, hk)
                    hpkMedia.append(hpk)

                    hk15.extend(hk_int_15)
                    hpk15 = tl.calcular_hpk(fluencias_acumuladas, energias_acumuladas, coeficientes_mutr_rho, hk15)
                    hpkMedia15.append(hpk15)

                    hk30.extend(hk_int_30)
                    hpk30 = tl.calcular_hpk(fluencias_acumuladas, energias_acumuladas, coeficientes_mutr_rho, hk30)
                    hpkMedia30.append(hpk30)

                    hk45.extend(hk_int_45)
                    hpk45 = tl.calcular_hpk(fluencias_acumuladas, energias_acumuladas, coeficientes_mutr_rho, hk45)
                    hpkMedia45.append(hpk45)

                    hk60.extend(hk_int_60)
                    hpk60 = tl.calcular_hpk(fluencias_acumuladas, energias_acumuladas, coeficientes_mutr_rho, hk60)
                    hpkMedia60.append(hpk60)

                    hk75.extend(hk_int_75)
                    hpk75 = tl.calcular_hpk(fluencias_acumuladas, energias_acumuladas, coeficientes_mutr_rho, hk75)
                    hpkMedia75.append(hpk75)

                np.set_printoptions(linewidth=np.inf)

                # MEAN ENERGY
                media_energia = np.mean(energiasMedia)
                desviacion_energia = np.std(energiasMedia, ddof=0)
                coef_variacion_energia = (desviacion_energia / media_energia) * 100

                # MEAN Kair
                media_kerma = np.mean(kermasMedia)
                desviacion_kerma = np.std(kermasMedia, ddof=0)
                coef_variacion_kerma = (desviacion_kerma / media_kerma) * 100

                # MEAN 1st HVL
                mean_hvl = np.mean(hvlmean)
                sd_hvl = np.std(hvlmean, ddof=0)
                v_hvl = (sd_hvl / mean_hvl) * 100

                mean_hvlCu = np.mean(hvlCumean)
                sd_hvlCu = np.std(hvlCumean, ddof=0)
                v_hvlCu = (sd_hvlCu / mean_hvlCu) * 100

                # MEAN 2nd HVL
                mean_hvl2 = np.mean(hvl2mean)
                sd_hvl2 = np.std(hvl2mean, ddof=0)
                v_hvl2 = (sd_hvl2 / mean_hvl2) * 100

                mean_hvl2Cu = np.mean(hvl2Cumean)
                sd_hvl2Cu = np.std(hvl2Cumean, ddof=0)
                v_hvl2Cu = (sd_hvl2Cu / mean_hvl2Cu) * 100

                # MEAN CONVERSION COEFFICIENTS AT DIFFERENT ANGLES
                media_hpk = np.mean(hpkMedia)
                media_hpk15 = np.mean(hpkMedia15)
                media_hpk30 = np.mean(hpkMedia30)
                media_hpk45 = np.mean(hpkMedia45)
                media_hpk60 = np.mean(hpkMedia60)
                media_hpk75 = np.mean(hpkMedia75)

                sd_hpk = np.std(hpkMedia, ddof=0)
                sd_hpk_15 = np.std(hpkMedia15, ddof=0)
                sd_hpk_30 = np.std(hpkMedia30, ddof=0)
                sd_hpk_45 = np.std(hpkMedia45, ddof=0)
                sd_hpk_60 = np.std(hpkMedia60, ddof=0)
                sd_hpk_75 = np.std(hpkMedia75, ddof=0)

                v_hpk = (sd_hpk / media_hpk) * 100
                v_hpk_15 = (sd_hpk_15 / media_hpk15) * 100
                v_hpk_30 = (sd_hpk_30 / media_hpk30) * 100
                v_hpk_45 = (sd_hpk_45 / media_hpk45) * 100
                v_hpk_60 = (sd_hpk_75 / media_hpk60) * 100
                v_hpk_75 = (sd_hpk_75 / media_hpk75) * 100

                tiempo_final_columns_7 = time()

                tiempo_ejecucion_columns_7 = tiempo_final_columns_7 - tiempo_inicial_columns_7

                # XCB: OUTPUT DATA DIGEST
                # ------------------------------------------------------------------------------------------------------
                # Lista de ángulos y sus respectivos valores hpk y sd_hpk
                angulos = [(0, float(hpk), sd_hpk, v_hpk), (15, float(hpk15), sd_hpk_15, v_hpk_15),
                           (30, float(hpk30), sd_hpk_30, v_hpk_30), (45, float(hpk45), sd_hpk_45, v_hpk_45),
                           (60, float(hpk60), sd_hpk_60, v_hpk_60), (75, float(hpk75), sd_hpk_75, v_hpk_75)]

                # List of magnitudes independent of the angle
                magnitudes = [(media_energia, desviacion_energia, coef_variacion_energia),
                              (media_kerma, desviacion_kerma, coef_variacion_kerma),
                              (mean_hvl, sd_hvl, v_hvl),
                              (mean_hvl2, sd_hvl2, v_hvl2),
                              (mean_hvlCu, sd_hvlCu, v_hvlCu),
                              (mean_hvl2Cu, sd_hvl2Cu, v_hvl2Cu)]

                # Save results in TXT and CSV files
                df = tl.output_digest(operational_magnitude=f_m, quality=calidad, mean_magnitudes=magnitudes,
                                      conversion_coefficients=angulos, execution_time=tiempo_ejecucion_columns_7,
                                      output_folder=ruta)

            # THIRD LOOP: monoenergetic conversion coefficient at 7 incident angles: 0 - 90 deg
            # ("hp_3_cyl.csv")
            # ----------------------------------------------------------------------------------------------------------
            elif len(hk_table.columns) == 8:
                print('THIRD LOOP: monoenergetic conversion coefficient at 7 incident angles: 0 - 90 deg')
                tiempo_inicial_columns_8 = time()

                # SLICING THE TABLES WHEN DISCOVERING ZEROES.THIS AVOIDS CRASHING WHEN TAKING LOGS AND DOING AKIMA
                hk_table15 = hk_table[hk_table['15angle'] > 0]
                hk_table30 = hk_table[hk_table['30angle'] > 0]
                hk_table45 = hk_table[hk_table['45angle'] > 0]
                hk_table60 = hk_table[hk_table['60angle'] > 0]
                hk_table75 = hk_table[hk_table['75angle'] > 0]
                hk_table90 = hk_table[hk_table['90angle'] > 0]

                Ehk = hk_table.iloc[:, 0].values
                hk_0 = hk_table.iloc[:, 1].values

                Ehk15 = hk_table15.iloc[:, 0].values
                hk_15 = hk_table15.iloc[:, 2].values

                Ehk30 = hk_table30.iloc[:, 0].values
                hk_30 = hk_table30.iloc[:, 3].values

                Ehk45 = hk_table45.iloc[:, 0].values
                hk_45 = hk_table45.iloc[:, 4].values

                Ehk60 = hk_table60.iloc[:, 0].values
                hk_60 = hk_table60.iloc[:, 5].values

                Ehk75 = hk_table75.iloc[:, 0].values
                hk_75 = hk_table75.iloc[:, 6].values

                Ehk90 = hk_table90.iloc[:, 0].values
                hk_90 = hk_table90.iloc[:, 7].values

                # TAKE LOGS FOR FURTHER AKIMA INTERPOLATIONS
                LEhk = []
                LEhk15 = []
                LEhk30 = []
                LEhk45 = []
                LEhk60 = []
                LEhk75 = []
                LEhk90 = []
                Lhk_0 = []
                Lhk_15 = []
                Lhk_30 = []
                Lhk_45 = []
                Lhk_60 = []
                Lhk_75 = []
                Lhk_90 = []

                [LEhk.append(math.log(x)) for x in Ehk]
                [LEhk15.append(math.log(x)) for x in Ehk15]
                [LEhk30.append(math.log(x)) for x in Ehk30]
                [LEhk45.append(math.log(x)) for x in Ehk45]
                [LEhk60.append(math.log(x)) for x in Ehk60]
                [LEhk75.append(math.log(x)) for x in Ehk75]
                [LEhk90.append(math.log(x)) for x in Ehk90]

                [Lhk_0.append(math.log(x)) for x in hk_0]
                [Lhk_15.append(math.log(x)) for x in hk_15]
                [Lhk_30.append(math.log(x)) for x in hk_30]
                [Lhk_45.append(math.log(x)) for x in hk_45]
                [Lhk_60.append(math.log(x)) for x in hk_60]
                [Lhk_75.append(math.log(x)) for x in hk_75]
                [Lhk_90.append(math.log(x)) for x in hk_90]

                # READ MUTR_RHO file
                mutr_rho = pd.read_csv(transmission_coefficients_file, sep="\t", header=None, encoding='ISO-8859-1')

                # Save variables
                E_mut_aux = mutr_rho.iloc[:, 0].values
                mtr_aux = mutr_rho.iloc[:, 1].values

                E_mut = []
                mtr = []

                # Replace dots by commas
                for x in E_mut_aux:
                    E_mut.append(float(x))

                for x in mtr_aux:
                    mtr.append(float(x))

                # Take logs
                LE_mut = [math.log(x) for x in E_mut]
                Lmtr = [math.log(x) for x in mtr]

                # NINI= Number of times that a random spectrum is generated (the parameters that define
                # a quality are changed around a central value using a gaussian or uniform distribution
                for j in range(nini):
                    if filter_Al != 0 and uAl != 0:
                        filter_Al_rand = np.random.uniform(lower_al, high_al, 1)
                    elif filter_Al != 0 and uAl == 0 and j == 0:
                        filter_Al_rand = np.random.uniform(lower_al, high_al, 1)
                    elif filter_Al == 0 and j == 0:
                        filter_Al_rand = np.random.uniform(lower_al, high_al, 1)

                    if filter_Be != 0 and uBe != 0:
                        filter_Be_rand = np.random.uniform(lower_be, high_be, 1)
                    elif filter_Be != 0 and uBe == 0 and j == 0:
                        filter_Be_rand = np.random.uniform(lower_be, high_be, 1)
                    elif filter_Be == 0 and j == 0:
                        filter_Be_rand = np.random.uniform(lower_be, high_be, 1)

                    if filter_Cu != 0 and uCu != 0:
                        filter_Cu_rand = np.random.uniform(lower_cu, high_cu, 1)
                    elif filter_Cu != 0 and uCu == 0 and j == 0:
                        filter_Cu_rand = np.random.uniform(lower_cu, high_cu, 1)
                    elif filter_Cu == 0 and j == 0:
                        filter_Cu_rand = np.random.uniform(lower_cu, high_cu, 1)

                    if filter_Sn != 0 and uSn != 0:
                        filter_Sn_rand = np.random.uniform(lower_sn, high_sn, 1)
                    elif filter_Sn != 0 and uSn == 0 and j == 0:
                        filter_Sn_rand = np.random.uniform(lower_sn, high_sn, 1)
                    elif filter_Sn == 0 and j == 0:
                        filter_Sn_rand = np.random.uniform(lower_sn, high_sn, 1)

                    if filter_Pb != 0 and uPb != 0:
                        filter_Pb_rand = np.random.uniform(lower_pb, high_pb, 1)
                    elif filter_Pb != 0 and uPb == 0 and j == 0:
                        filter_Pb_rand = np.random.uniform(lower_pb, high_pb, 1)
                    elif filter_Pb == 0 and j == 0:
                        filter_Pb_rand = np.random.uniform(lower_pb, high_pb, 1)

                    if filter_Air != 0 and uAir != 0:
                        filter_Air_rand = np.random.normal(filter_Air, uAir * filter_Air, 1)
                    elif filter_Air != 0 and uAir == 0 and j == 0:
                        filter_Air_rand = np.random.normal(filter_Air, uAir * filter_Air, 1)
                    elif filter_Air == 0 and j == 0:
                        filter_Air_rand = np.random.normal(filter_Air, uAir * filter_Air, 1)

                    if kvp != 0 and ukvp != 0:
                        kvp_rand = np.random.normal(kvp, ukvp * kvp, 1)
                    elif kvp != 0 and ukvp == 0 and j == 0:
                        kvp_rand = np.random.normal(kvp, ukvp * kvp, 1)
                    elif kvp == 0 and j == 0:
                        kvp_rand = np.random.normal(kvp, ukvp * kvp, 1)

                    if th != 0 and uth != 0:
                        th_rand = np.random.normal(th, uth * th, 1)
                    elif th != 0 and uth == 0 and j == 0:
                        th_rand = np.random.normal(th, uth * th, 1)
                    elif th_rand == 0 and j == 0:
                        th_rand = np.random.normal(th, uth * th, 1)

                    # SPEKPY
                    my_filters = [
                        ("Al", filter_Al_rand[0]),
                        ("Be", filter_Be_rand[0]),
                        ("Air", filter_Air_rand[0]),
                        ("Cu", filter_Cu_rand[0]),
                        ("Sn", filter_Sn_rand[0]),
                        ("Pb", filter_Pb_rand[0]),
                    ]

                    s = sp.Spek(kvp=kvp_rand[0], th=th_rand[0])
                    s.multi_filter(my_filters)

                    E, fluencia = s.get_spectrum(edges=False)  # Obtener el espectro
                    hvl = s.get_hvl1()  # Get the 1st half-value-layer mm Al
                    hvlmean.append(hvl)
                    hvl2 = s.get_hvl2()  # Get the 2nd half-value-layer mm Al
                    hvl2mean.append(hvl2)

                    hvlCu = s.get_hvl1(matl='Cu')  # Get the 1st half-value-layer mm Cu
                    hvl2Cu = s.get_hvl2(matl='Cu')  # Get the 2nd half-value-layer mm Cu
                    hvlCumean.append(hvlCu)
                    hvl2Cumean.append(hvl2Cu)

                    # SPEKPY
                    # FILTRADO DE SPEKPY Y OPERACIONES SOBRE ESPECTRO
                    mascara = E >= 8  # cortamos para que no de error al interpolar los hK monoenergetico (comienzan a veces en 7keV)
                    E_filtrado_temp = E[mascara]
                    fluencia_filtrada_temp = fluencia[mascara]

                    # Calculo de energia media
                    # primero debemos crear 2 listas vacias para que vaya acumuludando todos los datos de todas las iteraciones del bucle
                    fluencias_acumuladas = []
                    energias_acumuladas = []

                    # Agrego valores a las listas vacias
                    fluencias_acumuladas.extend(fluencia_filtrada_temp)
                    energias_acumuladas.extend(E_filtrado_temp)

                    energia_media = tl.calcular_energia_media(fluencias_acumuladas, energias_acumuladas)

                    energiasMedia.append(energia_media)

                    # Realizar EL LOGARITMO
                    LE = []
                    for x in E_filtrado_temp:
                        LE.append(math.log(x))

                    # Making Akima interpolation with MUTR_RHO (after taking logs)
                    interpolacion_uno_mut = interpolate.Akima1DInterpolator(LE_mut, Lmtr, axis=0)
                    list_mutr = []

                    for i in LE:
                        mutrrho = math.exp(interpolacion_uno_mut(i))
                        mutr = np.random.normal(mutrrho, umutrr * mutrrho,
                                                1)  # HERE mutr/rho is sampled randomly from gaussian distrib
                        list_mutr.append(mutr)

                    coeficientes_mutr_rho = []
                    coeficientes_mutr_rho.extend(list_mutr)

                    # Calculo de kerma medio fuera del bucle
                    kerma_medio = tl.calcular_kerma_medio(fluencias_acumuladas, energias_acumuladas,
                                                          coeficientes_mutr_rho)
                    kermasMedia.append(kerma_medio)

                    # REALIZAR INTERPOLACIÓN CON HK
                    interpolacion_uno_0 = interpolate.Akima1DInterpolator(LEhk, Lhk_0, axis=0)
                    interpolacion_uno_15 = interpolate.Akima1DInterpolator(LEhk15, Lhk_15, axis=0)
                    interpolacion_uno_30 = interpolate.Akima1DInterpolator(LEhk30, Lhk_30, axis=0)
                    interpolacion_uno_45 = interpolate.Akima1DInterpolator(LEhk45, Lhk_45, axis=0)
                    interpolacion_uno_60 = interpolate.Akima1DInterpolator(LEhk60, Lhk_60, axis=0)
                    interpolacion_uno_75 = interpolate.Akima1DInterpolator(LEhk75, Lhk_75, axis=0)
                    interpolacion_uno_90 = interpolate.Akima1DInterpolator(LEhk90, Lhk_90, axis=0)

                    ln_hk_int_0 = []
                    ln_hk_int_15 = []
                    ln_hk_int_30 = []
                    ln_hk_int_45 = []
                    ln_hk_int_60 = []
                    ln_hk_int_75 = []
                    ln_hk_int_90 = []

                    for i in LE:
                        interpolacion_final_0 = math.exp(interpolacion_uno_0(i))
                        ln_hk_int_0.append(interpolacion_final_0)

                        interpolacion_final_15 = math.exp(interpolacion_uno_15(i))
                        ln_hk_int_15.append(interpolacion_final_15)

                        interpolacion_final_30 = math.exp(interpolacion_uno_30(i))
                        ln_hk_int_30.append(interpolacion_final_30)

                        interpolacion_final_45 = math.exp(interpolacion_uno_45(i))
                        ln_hk_int_45.append(interpolacion_final_45)

                        interpolacion_final_60 = math.exp(interpolacion_uno_60(i))
                        ln_hk_int_60.append(interpolacion_final_60)

                        interpolacion_final_75 = math.exp(interpolacion_uno_75(i))
                        ln_hk_int_75.append(interpolacion_final_75)

                        interpolacion_final_90 = math.exp(interpolacion_uno_90(i))
                        ln_hk_int_90.append(interpolacion_final_90)

                    hk_int_0 = ln_hk_int_0
                    hk_int_15 = ln_hk_int_15
                    hk_int_30 = ln_hk_int_30
                    hk_int_45 = ln_hk_int_45
                    hk_int_60 = ln_hk_int_60
                    hk_int_75 = ln_hk_int_75
                    hk_int_90 = ln_hk_int_90

                    hk = []
                    hk15 = []
                    hk30 = []
                    hk45 = []
                    hk60 = []
                    hk75 = []
                    hk90 = []

                    hk.extend(hk_int_0)
                    hpk = tl.calcular_hpk(fluencias_acumuladas, energias_acumuladas, coeficientes_mutr_rho, hk)
                    hpkMedia.append(hpk)

                    hk15.extend(hk_int_15)
                    hpk15 = tl.calcular_hpk(fluencias_acumuladas, energias_acumuladas, coeficientes_mutr_rho, hk15)
                    hpkMedia15.append(hpk15)

                    hk30.extend(hk_int_30)
                    hpk30 = tl.calcular_hpk(fluencias_acumuladas, energias_acumuladas, coeficientes_mutr_rho, hk30)
                    hpkMedia30.append(hpk30)

                    hk45.extend(hk_int_45)
                    hpk45 = tl.calcular_hpk(fluencias_acumuladas, energias_acumuladas, coeficientes_mutr_rho, hk45)
                    hpkMedia45.append(hpk45)

                    hk60.extend(hk_int_60)
                    hpk60 = tl.calcular_hpk(fluencias_acumuladas, energias_acumuladas, coeficientes_mutr_rho, hk60)
                    hpkMedia60.append(hpk60)

                    hk75.extend(hk_int_75)
                    hpk75 = tl.calcular_hpk(fluencias_acumuladas, energias_acumuladas, coeficientes_mutr_rho, hk75)
                    hpkMedia75.append(hpk75)

                    hk90.extend(hk_int_90)
                    hpk90 = tl.calcular_hpk(fluencias_acumuladas, energias_acumuladas, coeficientes_mutr_rho, hk90)
                    hpkMedia90.append(hpk90)

                np.set_printoptions(linewidth=np.inf)

                # MEAN ENERGY
                media_energia = np.mean(energiasMedia)
                desviacion_energia = np.std(energiasMedia, ddof=0)
                coef_variacion_energia = (desviacion_energia / media_energia) * 100

                # MEAN Kair
                media_kerma = np.mean(kermasMedia)
                desviacion_kerma = np.std(kermasMedia, ddof=0)
                coef_variacion_kerma = (desviacion_kerma / media_kerma) * 100

                # MEAN 1st HVL
                mean_hvl = np.mean(hvlmean)
                sd_hvl = np.std(hvlmean, ddof=0)
                v_hvl = (sd_hvl / mean_hvl) * 100

                mean_hvlCu = np.mean(hvlCumean)
                sd_hvlCu = np.std(hvlCumean, ddof=0)
                v_hvlCu = (sd_hvlCu / mean_hvlCu) * 100

                # MEAN 2nd HVL
                mean_hvl2 = np.mean(hvl2mean)
                sd_hvl2 = np.std(hvl2mean, ddof=0)
                v_hvl2 = (sd_hvl2 / mean_hvl2) * 100

                mean_hvl2Cu = np.mean(hvl2Cumean)
                sd_hvl2Cu = np.std(hvl2Cumean, ddof=0)
                v_hvl2Cu = (sd_hvl2Cu / mean_hvl2Cu) * 100

                # MEDIA DEL ESPECTRO
                media_hpk = np.mean(hpkMedia)
                media_hpk15 = np.mean(hpkMedia15)
                media_hpk30 = np.mean(hpkMedia30)
                media_hpk45 = np.mean(hpkMedia45)
                media_hpk60 = np.mean(hpkMedia60)
                media_hpk75 = np.mean(hpkMedia75)
                media_hpk90 = np.mean(hpkMedia90)

                sd_hpk = np.std(hpkMedia, ddof=0)
                sd_hpk_15 = np.std(hpkMedia15, ddof=0)
                sd_hpk_30 = np.std(hpkMedia30, ddof=0)
                sd_hpk_45 = np.std(hpkMedia45, ddof=0)
                sd_hpk_60 = np.std(hpkMedia60, ddof=0)
                sd_hpk_75 = np.std(hpkMedia75, ddof=0)
                sd_hpk_90 = np.std(hpkMedia90, ddof=0)

                v_hpk = (sd_hpk / media_hpk) * 100
                v_hpk_15 = (sd_hpk_15 / media_hpk15) * 100
                v_hpk_30 = (sd_hpk_30 / media_hpk30) * 100
                v_hpk_45 = (sd_hpk_45 / media_hpk45) * 100
                v_hpk_60 = (sd_hpk_75 / media_hpk60) * 100
                v_hpk_75 = (sd_hpk_75 / media_hpk75) * 100
                v_hpk_90 = (sd_hpk_90 / media_hpk90) * 100

                tiempo_final_columns_8 = time()

                tiempo_ejecucion_columns_8 = tiempo_final_columns_8 - tiempo_inicial_columns_8

                # XCB: OUTPUT DATA DIGEST
                # ------------------------------------------------------------------------------------------------------
                # Lista de ángulos y sus respectivos valores hpk y sd_hpk
                angulos = [(0, float(hpk), sd_hpk, v_hpk), (15, float(hpk15), sd_hpk_15, v_hpk_15),
                           (30, float(hpk30), sd_hpk_30, v_hpk_30), (45, float(hpk45), sd_hpk_45, v_hpk_45),
                           (60, float(hpk60), sd_hpk_60, v_hpk_60), (75, float(hpk75), sd_hpk_75, v_hpk_75),
                           (90, float(hpk90), sd_hpk_90, v_hpk_90)]

                # List of magnitudes independent of the angle
                magnitudes = [(media_energia, desviacion_energia, coef_variacion_energia),
                              (media_kerma, desviacion_kerma, coef_variacion_kerma),
                              (mean_hvl, sd_hvl, v_hvl),
                              (mean_hvl2, sd_hvl2, v_hvl2),
                              (mean_hvlCu, sd_hvlCu, v_hvlCu),
                              (mean_hvl2Cu, sd_hvl2Cu, v_hvl2Cu)]

                # Save results in TXT and CSV files
                df = tl.output_digest(operational_magnitude=f_m, quality=calidad, mean_magnitudes=magnitudes,
                                      conversion_coefficients=angulos, execution_time=tiempo_ejecucion_columns_8,
                                      output_folder=ruta)

                print('El tiempo de ejecucion para hktable columns 8 en (s) fue:', tiempo_ejecucion_columns_8)

            # CUARTO LOOP: monoenergetic conversion coefficients at 8 incident angles: 0 - 180 deg
            # ("h_prime_3.csv", "h_prime_0.07.csv")
            # ----------------------------------------------------------------------------------------------------------
            elif len(hk_table.columns) == 9:
                print('CUARTO LOOP: monoenergetic conversion coefficients at 8 incident angles: 0 - 180 deg')
                tiempo_inicial_columns_9 = time()

                # SLICING THE TABLES WHEN DISCOVERING ZEROES.THIS AVOIDS CRASHING WHEN TAKING LOGS AND DOING AKIMA
                hk_table15 = hk_table[hk_table['15angle'] > 0]
                hk_table30 = hk_table[hk_table['30angle'] > 0]
                hk_table45 = hk_table[hk_table['45angle'] > 0]
                hk_table60 = hk_table[hk_table['60angle'] > 0]
                hk_table75 = hk_table[hk_table['75angle'] > 0]
                hk_table90 = hk_table[hk_table['90angle'] > 0]
                hk_table180 = hk_table[hk_table['180angle'] > 0]

                Ehk = hk_table.iloc[:, 0].values
                hk_0 = hk_table.iloc[:, 1].values

                Ehk15 = hk_table15.iloc[:, 0].values
                hk_15 = hk_table15.iloc[:, 2].values

                Ehk30 = hk_table30.iloc[:, 0].values
                hk_30 = hk_table30.iloc[:, 3].values

                Ehk45 = hk_table45.iloc[:, 0].values
                hk_45 = hk_table45.iloc[:, 4].values

                Ehk60 = hk_table60.iloc[:, 0].values
                hk_60 = hk_table60.iloc[:, 5].values

                Ehk75 = hk_table75.iloc[:, 0].values
                hk_75 = hk_table75.iloc[:, 6].values

                Ehk90 = hk_table90.iloc[:, 0].values
                hk_90 = hk_table90.iloc[:, 7].values

                Ehk180 = hk_table180.iloc[:, 0].values
                hk_180 = hk_table180.iloc[:, 8].values

                # TAKE LOGS FOR FURTHER AKIMA INTERPOLATIONS
                LEhk = []
                LEhk15 = []
                LEhk30 = []
                LEhk45 = []
                LEhk60 = []
                LEhk75 = []
                LEhk90 = []
                LEhk180 = []
                Lhk_0 = []
                Lhk_15 = []
                Lhk_30 = []
                Lhk_45 = []
                Lhk_60 = []
                Lhk_75 = []
                Lhk_90 = []
                Lhk_180 = []

                [LEhk.append(math.log(x)) for x in Ehk]
                [LEhk15.append(math.log(x)) for x in Ehk15]
                [LEhk30.append(math.log(x)) for x in Ehk30]
                [LEhk45.append(math.log(x)) for x in Ehk45]
                [LEhk60.append(math.log(x)) for x in Ehk60]
                [LEhk75.append(math.log(x)) for x in Ehk75]
                [LEhk90.append(math.log(x)) for x in Ehk90]
                [LEhk180.append(math.log(x)) for x in Ehk180]

                [Lhk_0.append(math.log(x)) for x in hk_0]
                [Lhk_15.append(math.log(x)) for x in hk_15]
                [Lhk_30.append(math.log(x)) for x in hk_30]
                [Lhk_45.append(math.log(x)) for x in hk_45]
                [Lhk_60.append(math.log(x)) for x in hk_60]
                [Lhk_75.append(math.log(x)) for x in hk_75]
                [Lhk_90.append(math.log(x)) for x in hk_90]
                [Lhk_180.append(math.log(x)) for x in hk_180]

                # READ MUTR_RHO file
                mutr_rho = pd.read_csv(transmission_coefficients_file, sep="\t", header=None, encoding='ISO-8859-1')

                # Save MUTR_RHO vs Energy
                E_mut_aux = mutr_rho.iloc[:, 0].values
                mtr_aux = mutr_rho.iloc[:, 1].values

                E_mut = []
                mtr = []

                # Replace dots by commas
                for x in E_mut_aux:
                    E_mut.append(float(x))

                for x in mtr_aux:
                    mtr.append(float(x))

                # Take logs
                LE_mut = [math.log(x) for x in E_mut]
                Lmtr = [math.log(x) for x in mtr]

                # NINI= Number of times that a random spectrum is generated (the parameters that define
                # a quality are changed around a central value using a gaussian or uniform distribution
                for j in range(nini):
                    if filter_Al != 0 and uAl != 0:
                        filter_Al_rand = np.random.uniform(lower_al, high_al, 1)
                    elif filter_Al != 0 and uAl == 0 and j == 0:
                        filter_Al_rand = np.random.uniform(lower_al, high_al, 1)
                    elif filter_Al == 0 and j == 0:
                        filter_Al_rand = np.random.uniform(lower_al, high_al, 1)

                    if filter_Be != 0 and uBe != 0:
                        filter_Be_rand = np.random.uniform(lower_be, high_be, 1)
                    elif filter_Be != 0 and uBe == 0 and j == 0:
                        filter_Be_rand = np.random.uniform(lower_be, high_be, 1)
                    elif filter_Be == 0 and j == 0:
                        filter_Be_rand = np.random.uniform(lower_be, high_be, 1)

                    if filter_Cu != 0 and uCu != 0:
                        filter_Cu_rand = np.random.uniform(lower_cu, high_cu, 1)
                    elif filter_Cu != 0 and uCu == 0 and j == 0:
                        filter_Cu_rand = np.random.uniform(lower_cu, high_cu, 1)
                    elif filter_Cu == 0 and j == 0:
                        filter_Cu_rand = np.random.uniform(lower_cu, high_cu, 1)

                    if filter_Sn != 0 and uSn != 0:
                        filter_Sn_rand = np.random.uniform(lower_sn, high_sn, 1)
                    elif filter_Sn != 0 and uSn == 0 and j == 0:
                        filter_Sn_rand = np.random.uniform(lower_sn, high_sn, 1)
                    elif filter_Sn == 0 and j == 0:
                        filter_Sn_rand = np.random.uniform(lower_sn, high_sn, 1)

                    if filter_Pb != 0 and uPb != 0:
                        filter_Pb_rand = np.random.uniform(lower_pb, high_pb, 1)
                    elif filter_Pb != 0 and uPb == 0 and j == 0:
                        filter_Pb_rand = np.random.uniform(lower_pb, high_pb, 1)
                    elif filter_Pb == 0 and j == 0:
                        filter_Pb_rand = np.random.uniform(lower_pb, high_pb, 1)

                    if filter_Air != 0 and uAir != 0:
                        filter_Air_rand = np.random.normal(filter_Air, uAir * filter_Air, 1)
                    elif filter_Air != 0 and uAir == 0 and j == 0:
                        filter_Air_rand = np.random.normal(filter_Air, uAir * filter_Air, 1)
                    elif filter_Air == 0 and j == 0:
                        filter_Air_rand = np.random.normal(filter_Air, uAir * filter_Air, 1)

                    if kvp != 0 and ukvp != 0:
                        kvp_rand = np.random.normal(kvp, ukvp * kvp, 1)
                    elif kvp != 0 and ukvp == 0 and j == 0:
                        kvp_rand = np.random.normal(kvp, ukvp * kvp, 1)
                    elif kvp == 0 and j == 0:
                        kvp_rand = np.random.normal(kvp, ukvp * kvp, 1)

                    if th != 0 and uth != 0:
                        th_rand = np.random.normal(th, uth * th, 1)
                    elif th != 0 and uth == 0 and j == 0:
                        th_rand = np.random.normal(th, uth * th, 1)
                    elif th_rand == 0 and j == 0:
                        th_rand = np.random.normal(th, uth * th, 1)

                    # SPEKPY
                    my_filters = [
                        ("Al", filter_Al_rand[0]),
                        ("Be", filter_Be_rand[0]),
                        ("Air", filter_Air_rand[0]),
                        ("Cu", filter_Cu_rand[0]),
                        ("Sn", filter_Sn_rand[0]),
                        ("Pb", filter_Pb_rand[0]),
                    ]

                    s = sp.Spek(kvp=kvp_rand[0], th=th_rand[0])
                    s.multi_filter(my_filters)

                    E, fluencia = s.get_spectrum(edges=False)  # Obtener el espectro
                    hvl = s.get_hvl1()  # Get the 1st half-value-layer mm Al
                    hvlmean.append(hvl)
                    hvl2 = s.get_hvl2()  # Get the 2nd half-value-layer mm Al
                    hvl2mean.append(hvl2)

                    hvlCu = s.get_hvl1(matl='Cu')  # Get the 1st half-value-layer mm Cu
                    hvl2Cu = s.get_hvl2(matl='Cu')  # Get the 2nd half-value-layer mm Cu
                    hvlCumean.append(hvlCu)
                    hvl2Cumean.append(hvl2Cu)
                    # SPEKPY

                    # FILTRADO DE SPEKPY Y OPERACIONES SOBRE ESPECTRO
                    mascara = E >= 8  # cortamos para que no de error al interpolar los hK monoenergetico (comienzan a veces en 7keV)
                    E_filtrado_temp = E[mascara]
                    fluencia_filtrada_temp = fluencia[mascara]

                    # Calculo de energia media
                    # primero debemos crear 2 listas vacias para que vaya acumuludando todos los datos de todas las iteraciones del bucle
                    fluencias_acumuladas = []
                    energias_acumuladas = []

                    # Agrego valores a las listas vacias
                    fluencias_acumuladas.extend(fluencia_filtrada_temp)
                    energias_acumuladas.extend(E_filtrado_temp)

                    energia_media = tl.calcular_energia_media(fluencias_acumuladas, energias_acumuladas)

                    energiasMedia.append(energia_media)

                    # Realizar EL LOGARITMO
                    LE = []
                    for x in E_filtrado_temp:
                        LE.append(math.log(x))

                    # Making Akima interpolation with MUTR_RHO (after taking logs)
                    interpolacion_uno_mut = interpolate.Akima1DInterpolator(LE_mut, Lmtr, axis=0)
                    list_mutr = []

                    for i in LE:
                        mutrrho = math.exp(interpolacion_uno_mut(i))
                        mutr = np.random.normal(mutrrho, umutrr * mutrrho,
                                                1)  # HERE mutr/rho is sampled randomly from gaussian distrib
                        list_mutr.append(mutr)

                    coeficientes_mutr_rho = []
                    coeficientes_mutr_rho.extend(list_mutr)

                    # Calculation of air kerma (out of the loop)
                    kerma_medio = tl.calcular_kerma_medio(fluencias_acumuladas, energias_acumuladas,
                                                          coeficientes_mutr_rho)
                    kermasMedia.append(kerma_medio)

                    # REALIZAR INTERPOLACIÓN CON HK
                    interpolacion_uno_0 = interpolate.Akima1DInterpolator(LEhk, Lhk_0, axis=0)
                    interpolacion_uno_15 = interpolate.Akima1DInterpolator(LEhk15, Lhk_15, axis=0)
                    interpolacion_uno_30 = interpolate.Akima1DInterpolator(LEhk30, Lhk_30, axis=0)
                    interpolacion_uno_45 = interpolate.Akima1DInterpolator(LEhk45, Lhk_45, axis=0)
                    interpolacion_uno_60 = interpolate.Akima1DInterpolator(LEhk60, Lhk_60, axis=0)
                    interpolacion_uno_75 = interpolate.Akima1DInterpolator(LEhk75, Lhk_75, axis=0)
                    interpolacion_uno_90 = interpolate.Akima1DInterpolator(LEhk90, Lhk_90, axis=0)
                    interpolacion_uno_180 = interpolate.Akima1DInterpolator(LEhk180, Lhk_180, axis=0)

                    ln_hk_int_0 = []
                    ln_hk_int_15 = []
                    ln_hk_int_30 = []
                    ln_hk_int_45 = []
                    ln_hk_int_60 = []
                    ln_hk_int_75 = []
                    ln_hk_int_90 = []
                    ln_hk_int_180 = []

                    for i in LE:
                        interpolacion_final_0 = math.exp(interpolacion_uno_0(i))
                        ln_hk_int_0.append(interpolacion_final_0)

                        interpolacion_final_15 = math.exp(interpolacion_uno_15(i))
                        ln_hk_int_15.append(interpolacion_final_15)

                        interpolacion_final_30 = math.exp(interpolacion_uno_30(i))
                        ln_hk_int_30.append(interpolacion_final_30)

                        interpolacion_final_45 = math.exp(interpolacion_uno_45(i))
                        ln_hk_int_45.append(interpolacion_final_45)

                        interpolacion_final_60 = math.exp(interpolacion_uno_60(i))
                        ln_hk_int_60.append(interpolacion_final_60)

                        interpolacion_final_75 = math.exp(interpolacion_uno_75(i))
                        ln_hk_int_75.append(interpolacion_final_75)

                        interpolacion_final_90 = math.exp(interpolacion_uno_90(i))
                        ln_hk_int_90.append(interpolacion_final_90)

                        interpolacion_final_180 = math.exp(interpolacion_uno_180(i))
                        ln_hk_int_180.append(interpolacion_final_180)

                    hk_int_0 = ln_hk_int_0
                    hk_int_15 = ln_hk_int_15
                    hk_int_30 = ln_hk_int_30
                    hk_int_45 = ln_hk_int_45
                    hk_int_60 = ln_hk_int_60
                    hk_int_75 = ln_hk_int_75
                    hk_int_90 = ln_hk_int_90
                    hk_int_180 = ln_hk_int_180

                    hk = []
                    hk15 = []
                    hk30 = []
                    hk45 = []
                    hk60 = []
                    hk75 = []
                    hk90 = []
                    hk180 = []

                    hk.extend(hk_int_0)
                    hpk = tl.calcular_hpk(fluencias_acumuladas, energias_acumuladas, coeficientes_mutr_rho, hk)
                    hpkMedia.append(hpk)

                    hk15.extend(hk_int_15)
                    hpk15 = tl.calcular_hpk(fluencias_acumuladas, energias_acumuladas, coeficientes_mutr_rho, hk15)
                    hpkMedia15.append(hpk15)

                    hk30.extend(hk_int_30)
                    hpk30 = tl.calcular_hpk(fluencias_acumuladas, energias_acumuladas, coeficientes_mutr_rho, hk30)
                    hpkMedia30.append(hpk30)

                    hk45.extend(hk_int_45)
                    hpk45 = tl.calcular_hpk(fluencias_acumuladas, energias_acumuladas, coeficientes_mutr_rho, hk45)
                    hpkMedia45.append(hpk45)

                    hk60.extend(hk_int_60)
                    hpk60 = tl.calcular_hpk(fluencias_acumuladas, energias_acumuladas, coeficientes_mutr_rho, hk60)
                    hpkMedia60.append(hpk60)

                    hk75.extend(hk_int_75)
                    hpk75 = tl.calcular_hpk(fluencias_acumuladas, energias_acumuladas, coeficientes_mutr_rho, hk75)
                    hpkMedia75.append(hpk75)

                    hk90.extend(hk_int_90)
                    hpk90 = tl.calcular_hpk(fluencias_acumuladas, energias_acumuladas, coeficientes_mutr_rho, hk90)
                    hpkMedia90.append(hpk90)

                    hk180.extend(hk_int_180)
                    hpk180 = tl.calcular_hpk(fluencias_acumuladas, energias_acumuladas, coeficientes_mutr_rho, hk180)

                    hpkMedia180.extend(hk180)

                np.set_printoptions(linewidth=np.inf)

                # MEAN ENERGY
                media_energia = np.mean(energiasMedia)
                desviacion_energia = np.std(energiasMedia, ddof=0)
                coef_variacion_energia = (desviacion_energia / media_energia) * 100

                # MEAN Kair
                media_kerma = np.mean(kermasMedia)
                desviacion_kerma = np.std(kermasMedia, ddof=0)
                coef_variacion_kerma = (desviacion_kerma / media_kerma) * 100

                # MEAN 1st HVL
                mean_hvl = np.mean(hvlmean)
                sd_hvl = np.std(hvlmean, ddof=0)
                v_hvl = (sd_hvl / mean_hvl) * 100

                mean_hvlCu = np.mean(hvlCumean)
                sd_hvlCu = np.std(hvlCumean, ddof=0)
                v_hvlCu = (sd_hvlCu / mean_hvlCu) * 100

                # MEAN 2nd HVL
                mean_hvl2 = np.mean(hvl2mean)
                sd_hvl2 = np.std(hvl2mean, ddof=0)
                v_hvl2 = (sd_hvl2 / mean_hvl2) * 100

                mean_hvl2Cu = np.mean(hvl2Cumean)
                sd_hvl2Cu = np.std(hvl2Cumean, ddof=0)
                v_hvl2Cu = (sd_hvl2Cu / mean_hvl2Cu) * 100

                # MEDIA DEL ESPECTRO
                media_hpk = np.mean(hpkMedia)
                media_hpk15 = np.mean(hpkMedia15)
                media_hpk30 = np.mean(hpkMedia30)
                media_hpk45 = np.mean(hpkMedia45)
                media_hpk60 = np.mean(hpkMedia60)
                media_hpk75 = np.mean(hpkMedia75)
                media_hpk90 = np.mean(hpkMedia90)
                media_hpk180 = np.mean(hpkMedia180)

                sd_hpk = np.std(hpkMedia, ddof=0)
                sd_hpk_15 = np.std(hpkMedia15, ddof=0)
                sd_hpk_30 = np.std(hpkMedia30, ddof=0)
                sd_hpk_45 = np.std(hpkMedia45, ddof=0)
                sd_hpk_60 = np.std(hpkMedia60, ddof=0)
                sd_hpk_75 = np.std(hpkMedia75, ddof=0)
                sd_hpk_90 = np.std(hpkMedia90, ddof=0)
                sd_hpk_180 = np.std(hpkMedia180, ddof=0)

                v_hpk = (sd_hpk / media_hpk) * 100
                v_hpk_15 = (sd_hpk_15 / media_hpk15) * 100
                v_hpk_30 = (sd_hpk_30 / media_hpk30) * 100
                v_hpk_45 = (sd_hpk_45 / media_hpk45) * 100
                v_hpk_60 = (sd_hpk_75 / media_hpk60) * 100
                v_hpk_75 = (sd_hpk_75 / media_hpk75) * 100
                v_hpk_90 = (sd_hpk_90 / media_hpk90) * 100
                v_hpk_180 = (sd_hpk_90 / media_hpk180) * 100

                # Compute execution time
                tiempo_final_columns_9 = time()

                tiempo_ejecucion_columns_9 = tiempo_final_columns_9 - tiempo_inicial_columns_9

                # XCB: OUTPUT DATA DIGEST
                # ------------------------------------------------------------------------------------------------------
                # Lista de ángulos y sus respectivos valores hpk y sd_hpk
                angulos = [(0, float(hpk), sd_hpk, v_hpk), (15, float(hpk15), sd_hpk_15, v_hpk_15),
                           (30, float(hpk30), sd_hpk_30, v_hpk_30), (45, float(hpk45), sd_hpk_45, v_hpk_45),
                           (60, float(hpk60), sd_hpk_60, v_hpk_60), (75, float(hpk75), sd_hpk_75, v_hpk_75),
                           (90, float(hpk90), sd_hpk_90, v_hpk_90), (180, float(hpk180), sd_hpk_180, v_hpk_180)]

                # List of magnitudes independent of the angle
                magnitudes = [(media_energia, desviacion_energia, coef_variacion_energia),
                              (media_kerma, desviacion_kerma, coef_variacion_kerma),
                              (mean_hvl, sd_hvl, v_hvl),
                              (mean_hvl2, sd_hvl2, v_hvl2),
                              (mean_hvlCu, sd_hvlCu, v_hvlCu),
                              (mean_hvl2Cu, sd_hvl2Cu, v_hvl2Cu)]

                # Save results in TXT and CSV files
                df = tl.output_digest(operational_magnitude=f_m, quality=calidad, mean_magnitudes=magnitudes,
                                      conversion_coefficients=angulos, execution_time=tiempo_ejecucion_columns_9,
                                      output_folder=ruta)

                print('El tiempo de ejecucion para hktable columns 9 en (s) fue:', tiempo_ejecucion_columns_9)
